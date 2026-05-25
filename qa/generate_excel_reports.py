# -*- coding: utf-8 -*-
"""Генерация двух Excel-файлов: тест-кейсы и тесты на ошибки."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent

HEADER_FILL = PatternFill("solid", fgColor="3D6B7A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
CELL_FONT = Font(size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def autosize(ws, max_width=55):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v).split("\n")[0]))
        ws.column_dimensions[letter].width = min(max_width, max(12, max_len + 2))


def style_header(ws, row=1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER


def style_body(ws, start_row=2):
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = CELL_FONT
            cell.alignment = WRAP
            cell.border = BORDER


def build_test_cases():
    rows = [
        [
            "TC-01",
            "Загрузка веб-интерфейса",
            "Front / HTTP",
            "Сервер Go запущен (порт 8080), каталог front доступен",
            "1. Открыть http://localhost:8080\n2. Дождаться полной загрузки страницы",
            "Отображаются шапка с логотипом, таблица, блок ориентации, карта, график; ошибок в консоли нет",
            "Высокий",
            "Пройден",
            "Страница открывается, элементы на месте",
            "2026-04-21",
            "Ручной прогон",
        ],
        [
            "TC-02",
            "Приём телеметрии по WebSocket",
            "Front / WS + Back",
            "Генератор gen.py подключён к TCP :9000, WS к /ws установлен",
            "1. Запустить back\n2. Запустить generator/gen.py\n3. Открыть UI",
            "В таблице появляются строки параметров; значения обновляются; график температур растёт",
            "Высокий",
            "Пройден",
            "Данные приходят стабильно по 2 с",
            "2026-04-21",
            "Ручной прогон",
        ],
        [
            "TC-03",
            "Отображение карты и трека",
            "Front / MapLibre",
            "Есть координаты в потоке телеметрии",
            "1. Убедиться, что маркер движется\n2. Проверить линию орбиты за несколько точек",
            "Маркер на карте соответствует lat/lon; линия трека видна поверх подложки",
            "Высокий",
            "Пройден",
            "После правки map.js трек отображается",
            "2026-04-21",
            "Ручной прогон",
        ],
        [
            "TC-04",
            "Экспорт PDF",
            "Back / API",
            "В памяти есть хотя бы один сеанс с кадрами телеметрии",
            "1. Нажать «Скачать PDF»\n2. Открыть файл",
            "Скачивается PDF; сеансы сгруппированы; есть таблица кадров (или сокращение + отсылка к CSV)",
            "Средний",
            "Пройден",
            "Файл telemetry_sessions.pdf формируется",
            "2026-04-21",
            "Ручной прогон",
        ],
        [
            "TC-05",
            "Экспорт CSV",
            "Back / API",
            "Есть накопленные сеансы",
            "1. Нажать «Скачать CSV»\n2. Открыть в редакторе / Excel",
            "UTF-8 с BOM; колонки session_id, времена, frame_index_in_session, поля телеметрии",
            "Средний",
            "Пройден",
            "Структура соответствует api/csv.go",
            "2026-04-21",
            "Ручной прогон",
        ],
        [
            "TC-06",
            "Смена языка интерфейса",
            "Front / i18n",
            "Страница загружена, есть данные в таблице",
            "1. Выбрать English\n2. Проверить заголовки таблицы и легенду графика\n3. Вернуть Русский",
            "«Параметр/Value» и подписи серий графика меняются согласно LANG",
            "Средний",
            "Пройден",
            "lang.js + renderTable + updateChartLocale работают согласованно",
            "2026-04-21",
            "Ручной прогон",
        ],
        [
            "TC-07",
            "Переключение тёмной темы",
            "Front / theme.js",
            "Страница открыта",
            "1. Нажать «Тёмная тема»\n2. Проверить фон, карточки, стиль карты\n3. Вернуть светлую",
            "Тема сохраняется в localStorage; карта переключает positron/dark-matter; график обновляет цвета осей",
            "Средний",
            "Пройден",
            "Переключение без перезагрузки страницы",
            "2026-04-21",
            "Ручной прогон",
        ],
        [
            "TC-08",
            "Сеансы связи (несколько TCP-подключений)",
            "Back / TCP + storage",
            "Сервер запущен",
            "1. Запустить генератор, остановить, запустить снова (новое подключение)\n2. Скачать CSV",
            "Разные session_id; в каждом сеансе несколько кадров",
            "Высокий",
            "Пройден",
            "Логика MemoryStore + события tcp соответствует ТЗ",
            "2026-04-21",
            "Ручной прогон",
        ],
        [
            "TC-09",
            "Кнопка центрирования карты",
            "Front / map.js",
            "Есть координаты спутника",
            "1. Сместить карту вручную\n2. Нажать кнопку центровки (левый верх)",
            "Карта плавно центрируется на текущей позиции",
            "Низкий",
            "Пройден",
            "easeTo отрабатывает",
            "2026-04-21",
            "Ручной прогон",
        ],
        [
            "TC-10",
            "Ориентация (Roll/Pitch/Yaw) на canvas",
            "Front / orientation.js",
            "Поток телеметрии активен",
            "1. Наблюдать блок ориентации\n2. Сравнить с числами в таблице",
            "Углы на canvas и в подписях согласованы с данными",
            "Средний",
            "Пройден",
            "Отображение чисел и цвет при превышении порога",
            "2026-04-21",
            "Ручной прогон",
        ],
    ]
    headers = [
        "ID",
        "Название",
        "Модуль",
        "Предусловия",
        "Шаги",
        "Ожидаемый результат",
        "Приоритет",
        "Результат прогона",
        "Фактический результат / комментарий",
        "Дата прогона",
        "Исполнитель / тип",
    ]
    return headers, rows


def build_error_tests():
    rows = [
        [
            "ERR-01",
            "Неверный бинарный кадр TCP",
            "Отправить в сокет неструктурированные байты (текст hello)",
            "Сервер логирует parse error; соединение не рвётся; следующие валидные кадры обрабатываются",
            "— (TCP)",
            "Пройден",
            "processor.Parse возвращает ошибку, main игнорирует кадр",
        ],
        [
            "ERR-02",
            "PDF при пустом хранилище",
            "GET /api/telemetry/pdf без накопленных сеансов",
            "HTTP 404, текст ошибки пользователю",
            "404",
            "Пройден",
            "TelemetryPDFHandler: «Нет данных телеметрии»",
        ],
        [
            "ERR-03",
            "CSV при пустом хранилище",
            "GET /api/telemetry/csv без данных",
            "HTTP 404",
            "404",
            "Пройден",
            "Аналогично pdf.go",
        ],
        [
            "ERR-04",
            "WebSocket недоступен",
            "Сервер остановлен, страница открыта",
            "WS не устанавливается; в консоли ошибка соединения; UI не падает",
            "—",
            "Пройден",
            "Ожидаемое поведение браузера",
        ],
        [
            "ERR-05",
            "Обрыв TCP во время сеанса",
            "Закрыть сокет генератора во время передачи",
            "Событие SessionClosed; сеанс получает EndedAt; UI продолжает работать",
            "—",
            "Пройден",
            "tcp handle defer шлёт RxSessionClosed",
        ],
        [
            "ERR-06",
            "Переполнение буфера канала событий (редко)",
            "Экстремальная частота пакетов",
            "Приёмник может блокироваться на отправке в канал (ограничение дизайна)",
            "—",
            "Блокирован / Не тестировался",
            "Документировать как ограничение; для курсовой достаточно штатной нагрузки",
        ],
        [
            "ERR-07",
            "Некорректный путь к шрифту PDF",
            "Удалить assets/fonts для теста",
            "Ошибка генерации PDF или падение при Output",
            "500 (возможно)",
            "Не выполнялся",
            "Негативный тест окружения; не запускался на чистой сборке",
        ],
        [
            "ERR-08",
            "Двойной клик по экспорту при отсутствии данных",
            "Хранилище пустое, пользователь жмёт PDF/CSV",
            "Каждый запрос возвращает 404",
            "404",
            "Пройден",
            "Повторяемость ответа сервера",
        ],
    ]
    headers = [
        "ID",
        "Негативный сценарий",
        "Условие / входные данные",
        "Ожидаемое поведение",
        "Ожидаемый код HTTP (если применимо)",
        "Результат прогона",
        "Комментарий / фактический итог",
    ]
    return headers, rows


def write_sheet(wb, title, headers, rows):
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws, 1)
    style_body(ws, 2)
    autosize(ws)


def main():
    wb1 = Workbook()
    write_sheet(wb1, "Тест-кейсы", *build_test_cases())
    p1 = OUT_DIR / "test_cases_results.xlsx"
    wb1.save(p1)
    print("Saved:", p1)

    wb2 = Workbook()
    # второй файл — новый workbook с другим именем листа
    ws2 = wb2.active
    ws2.title = "Тесты ошибок"
    h, r = build_error_tests()
    ws2.append(h)
    for row in r:
        ws2.append(row)
    style_header(ws2, 1)
    style_body(ws2, 2)
    autosize(ws2)
    p2 = OUT_DIR / "error_tests_results.xlsx"
    wb2.save(p2)
    print("Saved:", p2)


if __name__ == "__main__":
    main()
